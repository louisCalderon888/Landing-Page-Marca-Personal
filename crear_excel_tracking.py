#!/usr/bin/env python3
"""
Script para crear Excel de tracking de señales NQ con fórmulas automáticas
"""
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from datetime import datetime

# Crear nuevo workbook
wb = openpyxl.Workbook()
ws = wb.active
ws.title = "Señales NQ"

# Configurar anchos de columna
column_widths = {
    'A': 12,  # Fecha
    'B': 8,   # Hora
    'C': 8,   # Tipo
    'D': 15,  # Precio Entrada
    'E': 12,  # Stop Loss
    'F': 14,  # Take Profit
    'G': 15,  # Precio Salida
    'H': 18,  # Resultado (Puntos)
    'I': 12,  # R-Multiple
    'J': 12,  # Estado
    'K': 40   # Notas
}

for col, width in column_widths.items():
    ws.column_dimensions[col].width = width

# HEADERS (Fila 1)
headers = [
    'Fecha', 'Hora', 'Tipo', 'Precio Entrada', 'Stop Loss', 
    'Take Profit', 'Precio Salida', 'Resultado (Puntos)', 
    'R-Multiple', 'Estado', 'Notas'
]

# Estilos para headers
header_fill = PatternFill(start_color="0F4C81", end_color="0F4C81", fill_type="solid")
header_font = Font(bold=True, color="FFFFFF", size=11)
header_alignment = Alignment(horizontal="center", vertical="center")

for col_num, header in enumerate(headers, 1):
    cell = ws.cell(row=1, column=col_num)
    cell.value = header
    cell.fill = header_fill
    cell.font = header_font
    cell.alignment = header_alignment

# Agregar ejemplos con fórmulas (filas 2-4)
ejemplos = [
    ['2026-01-31', '09:30', 'LONG', 19500, 19450, 19600, 19600, '', '', 'GANANCIA', 'Ruptura de resistencia'],
    ['2026-01-30', '14:15', 'SHORT', 20100, 20150, 20000, 20000, '', '', 'GANANCIA', 'Rechazo en máximo'],
    ['2026-01-29', '10:00', 'LONG', 19800, 19750, 19900, '', '', '', 'PENDIENTE', 'Esperando cierre'],
]

for row_num, ejemplo in enumerate(ejemplos, 2):
    for col_num, value in enumerate(ejemplo, 1):
        cell = ws.cell(row=row_num, column=col_num)
        if col_num <= 7 or col_num >= 10:  # No fórmulas aún
            cell.value = value
        cell.alignment = Alignment(horizontal="center" if col_num <= 10 else "left", vertical="center")

# FÓRMULAS AUTOMÁTICAS
# Fórmula para Resultado (Puntos) - Columna H
for row in range(2, 100):  # Preparar 98 filas
    # H: Resultado (Puntos) = IF(Salida existe, IF(LONG, Salida-Entrada, Entrada-Salida), "")
    ws[f'H{row}'] = f'=IF(G{row}="","",IF(C{row}="LONG",G{row}-D{row},IF(C{row}="SHORT",D{row}-G{row},0)))'
    
    # I: R-Multiple = IF(Resultado existe, Resultado / ABS(Entrada - SL), "")
    ws[f'I{row}'] = f'=IF(H{row}="","",ROUND(H{row}/ABS(D{row}-E{row}),2)&"R")'

# Formatear números
for row in range(2, 100):
    for col in ['D', 'E', 'F', 'G']:
        ws[f'{col}{row}'].number_format = '#,##0.00'
    ws[f'H{row}'].number_format = '#,##0.00'

# DASHBOARD DE ESTADÍSTICAS (Columnas M-P)
stats_row = 2
stats_col = 13  # Columna M

# Títulos de estadísticas
stats_titles = [
    'ESTADÍSTICAS',
    'Total Señales:',
    'Señales Cerradas:',
    'Win Rate:',
    'Promedio R:',
    'Total Puntos:',
    'Max Ganancia:',
    'Max Pérdida:',
]

stats_formulas = [
    '',
    '=COUNTA(J2:J100)-COUNTIF(J2:J100,"")',
    '=COUNTIFS(J2:J100,"GANANCIA")+COUNTIFS(J2:J100,"PERDIDA")+COUNTIFS(J2:J100,"BREAKEVEN")',
    '=IF(N4=0,"0%",ROUND(COUNTIF(J2:J100,"GANANCIA")/N4*100,1)&"%")',
    '=IF(N4=0,0,ROUND(SUMPRODUCT((J2:J100="GANANCIA")+(J2:J100="PERDIDA")+(J2:J100="BREAKEVEN"),H2:H100)/N4,2))',
    '=SUMIFS(H2:H100,J2:J100,"GANANCIA")+SUMIFS(H2:H100,J2:J100,"PERDIDA")+SUMIFS(H2:H100,J2:J100,"BREAKEVEN")',
    '=IF(COUNT(H2:H100)=0,0,MAX(H2:H100))',
    '=IF(COUNT(H2:H100)=0,0,MIN(H2:H100))',
]

# Crear dashboard de estadísticas
for idx, (title, formula) in enumerate(zip(stats_titles, stats_formulas)):
    # Título en columna M
    title_cell = ws.cell(row=stats_row + idx, column=stats_col)
    title_cell.value = title
    title_cell.font = Font(bold=True, size=10)
    title_cell.alignment = Alignment(horizontal="right")
    
    # Valor/fórmula en columna N
    if formula:
        value_cell = ws.cell(row=stats_row + idx, column=stats_col + 1)
        value_cell.value = formula
        value_cell.font = Font(size=10, bold=True if idx == 0 else False)
        value_cell.alignment = Alignment(horizontal="left")

# Estilo para el título de estadísticas
ws.cell(row=stats_row, column=stats_col).font = Font(bold=True, size=12, color="0F4C81")
ws.cell(row=stats_row, column=stats_col + 1).font = Font(bold=True, size=12, color="0F4C81")

# Ajustar anchos de columnas de estadísticas
ws.column_dimensions['M'].width = 18
ws.column_dimensions['N'].width = 15

# Congelar primera fila
ws.freeze_panes = 'A2'

# Guardar archivo
filename = 'Tracking_Señales_NQ.xlsx'
wb.save(filename)
print(f"✅ Excel creado exitosamente: {filename}")
print(f"📊 Fórmulas automáticas configuradas:")
print(f"   - Resultado (Puntos): Se calcula automáticamente al ingresar Precio Salida")
print(f"   - R-Multiple: Se calcula automáticamente basado en el resultado")
print(f"   - Estadísticas: Win Rate, Promedio R, Total Puntos, etc.")
